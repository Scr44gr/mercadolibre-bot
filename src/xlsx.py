__author__ = "Scr44gr"
from typing import Any, Dict
import openpyxl
from typing import Union
from openpyxl.styles.fonts import Font
JSONType = Union[str, Any]

class ExcelFile:

    def __init__(self, filename: str) -> None:

        self._file = openpyxl.load_workbook(filename)
        self.current_sheet = None
        self.style = lambda **k: Font(**k)

    def select_sheet(self, name: str) -> Dict:

        sheets = self._file.sheetnames

        if name in sheets:
            self.current_sheet = self._file[name]
            return self.current_sheet
        raise SheetNotFound()
    
    @property
    def worksheet(self) -> openpyxl.worksheet.worksheet.Worksheet:
        return self._file

    def get_new_column_index(self):
        new_columns_index = self.current_sheet.max_column + 1
        return new_columns_index

class SheetNotFound(Exception):
    """
    The sheet is not found in the document.

    """